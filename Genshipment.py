# encoding='utf-8

# @Time: 2022-09-24
# @File: %
#!/usr/bin/env
from openpyxl import load_workbook
import pandas as pd
import polars as pl
from icecream import ic
import os
os.chdir(os.path.abspath(os.path.dirname(__file__)))
# change cwd to current file dir


def Read_files(df):
    dft = df.pivot(index=['SKU'], columns=['箱号'],
                   values='数量', aggregate_function='sum',)
    df_ununique = df.unique(maintain_order=True, subset=['箱号'], keep='first')
    df_lb = df_ununique.pivot(index=['SKU'], columns=['箱号'],
                              values='磅', aggregate_function='sum',).sum()
    df_size = df_ununique.pivot(index=['SKU'], columns=['箱号'],
                                values='箱型', aggregate_function='sum',).sum()
    df_lb[0, "SKU"] = '包装箱重量（磅）：'
    return [dft, df_lb, df_size]


def Gen_total_sku_NO(df):
    SKU_NO = df.groupby(['SKU']).agg([pl.col('数量').sum(),
                                      # pl.col('col_n').sum(),
                                      # pl.col('col_n').sum(),
                                      # pl.col('sum()').sum(),
                                      ]).sort(by='SKU', descending=True)

    return SKU_NO


def Gen_Box_info(df):
    df_number = df.groupby(['箱号', ]).agg([pl.col('数量').sum(),
                                          ]).sort(by='箱号', descending=False)
    df_weghit = df.unique(maintain_order=True, subset=[
                          '箱号'], keep='first')
    df_weghit = df_weghit.select([pl.col('箱号'),
                                  pl.col('磅'),
                                  ])
    df_merge = df_number.join(df_weghit, how='left', on='箱号', suffix='_x')
    return df_merge


# SKU_NO = Gen_total_sku_NO(df)
# ic(SKU_NO)

def Read_box_size(settingfile):
    df_box = pl.read_excel(settingfile, sheet_name='settings',
                           )
    cols = ['箱型', '包装箱宽度（英寸）：', '包装箱长度（英寸）：', '包装箱高度（英寸）：']
    df_box = df_box[cols]
    return df_box


def cast_str_to_float(data, col_name):
    return data.with_columns(pl.col(col_name).cast(pl.Float64, strict=False))


def Gen_Detail_package(filename, df_list, settingfile):
    dft = df_list[0]
    df_lb = df_list[1]
    df_size = df_list[2]
    # ic(df_size)
    df_package = pl.from_pandas(pd.read_excel(filename, sheet_name=0,
                                              header=0, skiprows=4,
                                              ))
    df_pacakge_sku = df_package.select([pl.col('SKU'),
                                        ])
    # join df
    df_joined = df_pacakge_sku.join(dft, how='left', on='SKU', suffix='_x')
    # df_joined.write_csv('dfjoined.csv')
    df_joined = df_joined[:-6]
    # df_joined.write_csv('df_joined_first.csv')
    # df_joined = pl.concat([df_joined, df_lb])
    # convert df_joined columns type from int to float
    for col in df_joined.columns[1:]:
        df_joined = cast_str_to_float(df_joined, col)
    # ic(df_joined.schema)
    # ic(df_lb.schema)

    df_joined = df_joined.vstack(df_lb)
    # df_joined.write_csv('df_joined.csv')

    # handle df_size
    col_name = df_joined.columns
    df_clone = df_joined[:3]
    df_clone[0, 'SKU'] = '包装箱宽度（英寸）：'
    df_clone[1, 'SKU'] = '包装箱长度（英寸）：'
    df_clone[2, 'SKU'] = '包装箱高度（英寸）：'
    box_size_info = Read_box_size(settingfile)
    # ic(box_size_info)
    for i in range(1, len(col_name)):
        box_int = df_size[col_name[i]][0]
        for j in range(box_size_info.height):
            if int(box_int) == int(box_size_info['箱型'][j]):
                df_clone[0, col_name[i]] = box_size_info['包装箱宽度（英寸）：'][j]
                df_clone[1, col_name[i]] = box_size_info['包装箱长度（英寸）：'][j]
                df_clone[2, col_name[i]] = box_size_info['包装箱高度（英寸）：'][j]
    # df_joined = pl.concat([df_joined, df_clone])
    df_joined = df_joined.vstack(df_clone)

    # df_joined.write_csv('df_joined.csv')

    return df_joined


def Write_package_info(templatefile, SKU_NO, outfilename):
    wb = load_workbook(templatefile)
    ws = wb['Create workflow – template']
    for i in range(SKU_NO.height):
        ws.cell(row=i + 9, column=1, value=SKU_NO['SKU'][i])
        ws.cell(row=i + 9, column=2, value=SKU_NO['数量'][i])
    wb.save(outfilename)
    return


def Write_box_info(templatefile, df_joined, outfilename):
    wb = load_workbook(templatefile)
    # df_joined.write_csv('dfjoined.csv')
    try:
        ws = wb.worksheets[0]
    except:
        ws = wb['Box packing information']
    # 先写数据
    for i in range(df_joined.height-4):
        for j in range(df_joined.width - 1):
            # print(i+6, j+13)
            ws.cell(row=i + 6, column=j + 13,
                    value=df_joined[df_joined.columns[j+1]][i])
    # 再写尺寸和重量
    for i in range(df_joined.height-4, df_joined.height):
        for j in range(df_joined.width - 1):
            ws.cell(row=i + 8, column=j + 13,
                    value=df_joined[df_joined.columns[j+1]][i])
    wb.save(outfilename)
    return


# df = pl.read_excel('装箱单.xlsx', sheet_name='template',
#                    # read_csv_options={'ignore_errors':True,'skip_rows':0,'dtypes':{} }
#                    )
# #
# df_list = Read_files(df)
# # ic(df_list[0])
# # ic(df_list[1])
# # ic(df_list[2])
# SKU_NO = Gen_total_sku_NO(df)
# print(SKU_NO.head(5))
# # Write_package_info('货件信息模板.xlsx', SKU_NO, '货件信息上传表-1.xlsx')
# #
# # df_joined = Gen_Detail_package('2022-10-21.xlsx', df_list, '装箱单.xlsx')
# # #
# # Write_box_info('2022-10-21.xlsx', df_joined, '2022-09-16包装信息_1.xlsx')
